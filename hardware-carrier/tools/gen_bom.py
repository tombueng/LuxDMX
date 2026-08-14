#!/usr/bin/env python3
"""Generate the bill of materials as Markdown, CSV and XLSX.

Quantities come from the board, sourcing comes from sourcing.json. Keeping them apart means
a moved part changes the count automatically and nobody has to remember to edit a table.

Two sections, because they are bought in different places:
  **LCSC/JLCPCB** for the small stuff (1206, terminals, polyfuse, electrolytic, buffer)
  **Amazon / eBay / AliExpress** for the modules (ESP32, MAX3485, W5500, buck, OLED, encoder)

Marketplace links rot within months. The search terms do not, so both are in the output and
the search terms are the part you should actually rely on.

Fails if a part on the board has no sourcing entry.

Run:  python hardware-carrier/gen_bom.py
"""
import csv
import json
import os
import re
import sys

import pcbnew

def _project_dir():
    """The project directory, whether this script sits in it or in tools/ beside it.

    The tooling is deliberately kept out of the git repo (see .gitignore), so it lives one
    level down in tools/. Everything it reads and writes still belongs next to the board."""
    d = os.path.dirname(os.path.abspath(__file__))
    for _ in range(3):
        if os.path.exists(os.path.join(d, "luxdmx-carrier.kicad_pcb")) or \
           os.path.exists(os.path.join(d, "modules.json")):
            return d
        d = os.path.dirname(d)
    return os.path.dirname(os.path.abspath(__file__))


HERE = _project_dir()
BOARD = os.path.join(HERE, "luxdmx-carrier.kicad_pcb")
# sourcing.json is tooling input, not a project file: it lives next to this script
SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sourcing.json")
OUT = os.path.join(HERE, "production")


def generic(val, fpid):
    """Board value -> sourcing key.

    Every 2.54 mm header is the same line item no matter what it is labelled, so the
    decision is made on the FOOTPRINT. The display and encoder labels sit on headers, not on
    the modules themselves: what you solder is a pin header, what you plug in is the module,
    and both have to be bought."""
    if "PinHeader" in fpid:
        return "PinHeader"
    if "header-1x03-P2.54-terminal" in fpid:
        # the DMX rows are drilled 1.20 so they take a screw terminal too, so they are their
        # own line, not just another piece off the header strip
        return "header-1x03-P2.54-terminal"
    v = re.sub(r"^(MAX3485 DMX)\d$", r"\1n", val)
    v = re.sub(r"^PIX\d (.*)", r"PIXn \1", v)
    if v.startswith("MF-R 30V"):
        return "MF-R 30V"
    if v.startswith("DC jack"):
        return "DC jack"
    return v


def main():
    src = json.load(open(SRC, encoding="utf-8"))
    parts, modules = src["parts"], src["modules"]
    # bewusst nicht in der Liste, aber trotzdem bekannt: sonst schlaegt das Gate unten an
    ignore = src.get("ignore", {})
    for k, v in parts.items():
        if "same_as" in v:
            parts[k] = dict(parts[v["same_as"]])

    board = pcbnew.LoadBoard(BOARD)
    counts, missing = {}, []
    for f in board.GetFootprints():
        val = f.GetValue()
        g = generic(val, f.GetFPIDAsString())
        key = g if (g in parts or g in modules or g in ignore) else \
            (val if val in modules else None)
        if key in ignore:
            continue
        if key is None:
            # headers are all one line item
            missing.append(val)
            continue
        counts[key] = counts.get(key, 0) + 1

    # The OLED and the encoder are not footprints, they plug into headers. One each.
    # The OLED is the only part left that is not a footprint of its own; it plugs into a
    # header. The encoder used to be listed the same way, as a KY-040 module on a 5-pin row.
    # It is a footprint now, an EC11 on the carrier, so it counts itself.
    for key, marker in (("OLED SDA/SCL/VCC/GND", "SDA/SCL"),):
        if any(marker in f.GetValue() for f in board.GetFootprints()):
            counts[key] = 1

    if missing:
        print("Ohne Bezugsquelle:")
        for m in sorted(set(missing)):
            print(f"   {m}")
        return 1

    # merge LCSC lines that are the same physical part (three 10k positions, one reel)
    merged = {}
    for key, n in sorted(counts.items()):
        if key not in parts:
            continue
        p = parts[key]
        # Without a part number there is nothing to merge ON: three marketplace parts all
        # answer (None, None) and came out as one line of four pieces reading
        # "fuse holder / EC11 / tact switch".
        mk = (p.get("mpn"), p.get("lcsc")) if (p.get("mpn") or p.get("lcsc")) else (key,)
        if mk in merged:
            merged[mk][1] += n
            merged[mk][2].append(key)
        else:
            merged[mk] = [p, n, [key]]

    rows_l, rows_m = [], []
    for p, n, names in sorted(merged.values(), key=lambda t: t[0]["desc"]):
        if True:
            key = " / ".join(sorted(set(names)))
            rows_l.append({
                "Stk": n, "Bauteil": key, "Beschreibung": p["desc"],
                "Hersteller": p.get("vendor", ""), "MPN": p.get("mpn", ""),
                "LCSC": p.get("lcsc") or "", "Bestand": p.get("stock") or "?",
                "EUR/Stk": p.get("price_eur", ""),
                "EUR gesamt": round(n * (p.get("price_eur") or 0), 2),
                "Link": p.get("url", ""), "Hinweis": p.get("note", ""),
                "_links": p.get("links"), "_source": p.get("source", "LCSC"),
                "_alt": p.get("alternatives"),
            })
    for key, n in sorted(counts.items()):
        if key in modules:
            m = modules[key]
            rows_m.append({
                "Stk": n, "Bauteil": key, "Beschreibung": m["desc"],
                "Suchbegriffe": " | ".join(f"{k}: {v}" for k, v in m["search"].items()),
                "Angebote": len(m["links"]), "Hinweis": m.get("note", ""),
                "_links": m["links"],
            })

    os.makedirs(OUT, exist_ok=True)
    total_l = sum(r["EUR gesamt"] for r in rows_l)

    # ---- Markdown ----------------------------------------------------------------
    L = ["# Bill of materials", "",
         f"Everything fitted: {sum(counts.values())} positions. Quantities come from the board.", "",
         f"LCSC stock checked **{src['_checked']}**. Look again before ordering, it changes "
         "daily.", "",
         src.get("_price_basis", ""), "",
         "## Small parts from LCSC", "",
         "| Qty | Part | Description | MPN | LCSC | Stock | EUR ea. | EUR total |",
         "|---:|---|---|---|---|---:|---:|---:|"]
    for r in rows_l:
        lc = f"[{r['LCSC']}]({r['Link']})" if r["LCSC"] and r["Link"] else (r["LCSC"] or "-")
        L.append(f"| {r['Stk']} | {r['Bauteil']} | {r['Beschreibung']} | `{r['MPN']}` | "
                 f"{lc} | {r['Bestand']} | {r['EUR/Stk']} | {r['EUR gesamt']} |")
    L += ["", f"**Small parts total: about {total_l:.2f} EUR**", ""]
    if ignore:
        L += ["Deliberately not listed:", ""]
        for k, why in ignore.items():
            L.append(f"- **{k}**: {why}")
        L.append("")
    notes = [r for r in rows_l if r["Hinweis"]]
    if notes:
        L += ["### Notes", ""]
        for r in notes:
            L.append(f"- **{r['Bauteil']}**: {r['Hinweis']}")
            # a Kleinteil can carry marketplace links too, when LCSC has nothing suitable
            for plat, title, url, price in r.get("_links") or []:
                L.append(f"  - {plat}: [{title}]({url}) - {price}")
        L.append("")
        for r in rows_l:
            if not r.get("_alt"):
                continue
            L += [f"### What may go in the {r['Bauteil']} position, and what may not", ""]
            for head, grp in r["_alt"].items():
                # a group is either a plain list of rows or {cols, rows}: the column headings
                # differ per table, a resistor ladder is not a chip family list
                cols = grp.get("cols") if isinstance(grp, dict) else None
                rows = grp["rows"] if isinstance(grp, dict) else grp
                cols = cols or ["Type", "", "", "Why"]
                L += [f"**{head}**", "",
                      "| " + " | ".join(cols) + " |",
                      "|" + "---|" * len(cols)]
                for row in rows:
                    L.append("| `" + str(row[0]) + "` | " +
                             " | ".join(str(x) for x in row[1:]) + " |")
                L.append("")

    L += ["## Modules from Amazon / eBay / AliExpress", "",
          "The links are examples from " + src["_checked"] +
          " and will rot. The **search terms** are the durable part.", ""]
    for r in rows_m:
        L += [f"### {r['Stk']} x {r['Bauteil']}", "", r["Beschreibung"], "",
              "| Platform | Search term |", "|---|---|"]
        for k, v in modules[r["Bauteil"]]["search"].items():
            L.append(f"| {k} | `{v}` |")
        L += ["", "| Platform | Offer | Price |", "|---|---|---|"]
        for plat, title, url, price in r["_links"]:
            L.append(f"| {plat} | [{title}]({url}) | {price} |")
        if r["Hinweis"]:
            L += ["", f"> {r['Hinweis']}"]
        L.append("")

    open(os.path.join(OUT, "BOM.md"), "w", encoding="utf-8").write("\n".join(L) + "\n")

    # ---- CSV ---------------------------------------------------------------------
    csv_rows = []
    for r in rows_l:
        csv_rows.append({"Group": "LCSC", "Qty": r["Stk"], "Part": r["Bauteil"],
                         "Description": r["Beschreibung"], "Manufacturer": r["Hersteller"],
                         "MPN": r["MPN"], "LCSC": r["LCSC"], "Stock": r["Bestand"],
                         "EUR ea": r["EUR/Stk"], "EUR total": r["EUR gesamt"],
                         "Source": r["Link"], "Note": r["Hinweis"]})
    for r in rows_l:
        for plat, title, url, price in r.get("_links") or []:
            csv_rows.append({"Group": plat, "Qty": r["Stk"], "Part": r["Bauteil"],
                             "Description": r["Beschreibung"], "Manufacturer": "",
                             "MPN": "", "LCSC": "", "Stock": "",
                             "EUR ea": price, "EUR total": "",
                             "Source": url, "Note": title})
    for r in rows_m:
        for plat, title, url, price in r["_links"]:
            csv_rows.append({"Group": plat, "Qty": r["Stk"], "Part": r["Bauteil"],
                             "Description": r["Beschreibung"], "Manufacturer": "",
                             "MPN": "", "LCSC": "", "Stock": "",
                             "EUR ea": price, "EUR total": "",
                             "Source": url, "Note": title})
    cols = ["Group", "Qty", "Part", "Description", "Manufacturer", "MPN", "LCSC",
            "Stock", "EUR ea", "EUR total", "Source", "Note"]
    with open(os.path.join(OUT, "BOM.csv"), "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(csv_rows)

    # ---- LCSC-Upload ---------------------------------------------------------------
    # https://www.lcsc.com/bom nimmt Menge + (LCSC-Nummer | MPN | Beschreibung). Teile ohne
    # bestaetigte C-Nummer gehen mit der MPN rein, die matcht LCSC selbst -- oder eben nicht,
    # dann stehen sie danach rot in der Liste. Besser als eine geratene Nummer.
    boards = int(sys.argv[sys.argv.index("--boards") + 1]) if "--boards" in sys.argv else 1
    up = []
    for r in rows_l:
        # only what LCSC can actually match; marketplace parts would just sit there red
        if r["_source"] != "LCSC":
            continue
        up.append({"LCSC Part Number": r["LCSC"].split(" / ")[0],
                   "Manufacturer Part Number": r["MPN"].split(" / ")[0],
                   "Quantity": r["Stk"] * boards,
                   "Designator": r["Bauteil"],
                   "Description": r["Beschreibung"]})
    with open(os.path.join(OUT, "BOM-LCSC-upload.csv"), "w", newline="",
              encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["LCSC Part Number", "Manufacturer Part Number",
                                           "Quantity", "Designator", "Description"])
        w.writeheader()
        w.writerows(up)

    # ---- XLSX --------------------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "LCSC"
        ws.append(cols[:-2] + ["Quelle", "Hinweis"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in csv_rows:
            if r["Group"] == "LCSC":
                ws.append([r[c] for c in cols])
        ws2 = wb.create_sheet("Module")
        ws2.append(["Platform", "Qty", "Part", "Offer", "Price", "Link"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for r in csv_rows:
            if r["Group"] != "LCSC":
                ws2.append([r["Group"], r["Qty"], r["Part"], r["Note"],
                            r["EUR ea"], r["Source"]])
        ws3 = wb.create_sheet("Search terms")
        ws3.append(["Part", "Platform", "Search term"])
        for c in ws3[1]:
            c.font = Font(bold=True)
        for r in rows_m:
            for k, v in modules[r["Bauteil"]]["search"].items():
                ws3.append([r["Bauteil"], k, v])
        for sheet in (ws, ws2, ws3):
            for col in sheet.columns:
                width = max(len(str(c.value or "")) for c in col)
                sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 60)
                for c in col:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(os.path.join(OUT, "BOM.xlsx"))
        xl = "BOM.xlsx"
    except ImportError:
        xl = "(openpyxl fehlt)"

    print(f"{len(rows_l)} LCSC-Positionen ({total_l:.2f} EUR), {len(rows_m)} Module")
    print(f"-> production/BOM.md, BOM.csv, {xl}, BOM-LCSC-upload.csv "
          f"({len(up)} Zeilen fuer {boards} Board(s))")
    return 0


if __name__ == "__main__":
    sys.exit(main())
